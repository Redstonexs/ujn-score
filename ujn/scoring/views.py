import io
import json
import logging
import os
import re
import time
import zipfile
from collections import defaultdict, Counter
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from threading import Lock

import qrcode
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook

from django.db import models
from .models import Category, Judge, Participant, Score, SiteConfig, Vote
from .sse import score_event_bus, build_scores_data

logger = logging.getLogger(__name__)

# Issue 6 & 16: 清空/管理操作密码从环境变量读取，不再硬编码
ADMIN_CLEAR_PASSWORD = os.getenv('CLEAR_PASSWORD', 'jndx')


# ==========================================================================
# 安全说明（Issue 20）:
# 本项目为前后端分离架构，前端通过 AJAX 调用后端 API。
# 使用 @csrf_exempt 是因为前端独立部署，无法使用 Django 的 CSRF Cookie 机制。
# 管理员接口通过密码验证保护，评委接口通过 UUID Token 保护。
# 生产环境建议：
#   1. 使用 HTTPS 加密传输
#   2. 配置 CORS 限制为前端域名
#   3. 如需更高安全性，可引入 JWT Token 认证替代密码验证
# ==========================================================================


# Issue 18: 简单的基于内存的限流装饰器
_rate_limit_store = {}
_rate_limit_lock = Lock()


def rate_limit(max_calls=10, period=60):
    """简单的基于 IP 的限流装饰器

    Args:
        max_calls: 时间窗口内允许的最大请求数
        period: 时间窗口（秒）
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() \
                or request.META.get('REMOTE_ADDR', 'unknown')
            key = f"{view_func.__name__}:{ip}"
            now = time.time()

            with _rate_limit_lock:
                if key in _rate_limit_store:
                    calls = _rate_limit_store[key]
                    # 清理过期记录
                    calls = [t for t in calls if now - t < period]
                    if len(calls) >= max_calls:
                        return error_response('请求过于频繁，请稍后再试', 429)
                    calls.append(now)
                    _rate_limit_store[key] = calls
                else:
                    _rate_limit_store[key] = [now]
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def json_response(data, status=200):
    return JsonResponse(data, status=status, json_dumps_params={'ensure_ascii': False})


def error_response(msg, status=400):
    return json_response({'error': msg}, status=status)


def parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on', 'y'}


def to_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def to_non_negative_int(value, default=0):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed >= 0 else default


def normalize_text(value):
    return str(value).strip() if value is not None else ''


def get_next_judge_order():
    max_order = Judge.objects.aggregate(models.Max('order'))['order__max']
    return (max_order or 0) + 1


def get_next_category_order():
    max_order = Category.objects.aggregate(models.Max('order'))['order_max']
    return (max_order or 0) + 1


def get_category_scoring_mode(category):
    """获取类别的实际打分模式"""
    if category.scoring_mode == 'default':
        config = SiteConfig.get_config()
        return config.scoring_mode
    return category.scoring_mode


def get_category_vote_params(category):
    """获取类别的投票参数"""
    config = SiteConfig.get_config()
    # 投票总数应该等于实际的选手人数
    participant_count = category.participants.count()
    if category.scoring_mode == 'default':
        select = config.vote_select_count
        return participant_count, select
    elif category.scoring_mode == 'vote':
        select = category.vote_select_count or config.vote_select_count
        return participant_count, select
    return None, None


def get_category_score_params(category):
    """获取类别的分数参数"""
    config = SiteConfig.get_config()
    if category.scoring_mode == 'default':
        return (
            config.score_min,
            config.score_max,
            config.score_value_type,
            config.allow_duplicate_scores,
            config.exclude_extreme_scores,
            config.exclude_lowest_count,
            config.exclude_highest_count,
        )
    elif category.scoring_mode == 'score':
        return (
            category.score_min or config.score_min,
            category.score_max or config.score_max,
            category.score_value_type or config.score_value_type,
            category.allow_duplicate_scores if category.allow_duplicate_scores is not None else config.allow_duplicate_scores,
            category.exclude_extreme_scores if category.exclude_extreme_scores is not None else config.exclude_extreme_scores,
            category.exclude_lowest_count if category.exclude_lowest_count is not None else config.exclude_lowest_count,
            category.exclude_highest_count if category.exclude_highest_count is not None else config.exclude_highest_count,
        )
    return (
        config.score_min,
        config.score_max,
        config.score_value_type,
        config.allow_duplicate_scores,
        config.exclude_extreme_scores,
        config.exclude_lowest_count,
        config.exclude_highest_count,
    )


def normalize_header(value):
    return normalize_text(value).replace(' ', '').lower()


def normalize_decimal(value, default=Decimal("0")):
    if value is None or value == '':
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return default


def quantize_score(value):
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def format_score_value(value):
    if value is None:
        return None
    decimal_value = quantize_score(normalize_decimal(value))
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)


def format_score_list(values):
    formatted = [str(format_score_value(value)) for value in values]
    return '、'.join(formatted)


def judge_display_name(judge_or_id):
    if hasattr(judge_or_id, 'name'):
        return judge_or_id.name
    judge_id = judge_or_id.id if hasattr(judge_or_id, 'id') else judge_or_id
    return f'评委{judge_id}'


def get_judge_allowed_category_ids(judge):
    return [category.id for category in judge.allowed_categories.all()]


def judge_can_access_category(judge, category):
    allowed_category_ids = get_judge_allowed_category_ids(judge)
    return not allowed_category_ids or category.id in allowed_category_ids


def serialize_judge(judge):
    allowed_categories = list(judge.allowed_categories.all())
    allowed_category_ids = [category.id for category in allowed_categories]
    return {
        'id': judge.id,
        'order': judge.order,
        'name': judge.name,
        'display_name': judge_display_name(judge),
        'token': str(judge.token),
        'is_active': judge.is_active,
        'allowed_category_ids': allowed_category_ids,
        'allowed_category_names': [category.name for category in allowed_categories],
        'all_categories_allowed': len(allowed_category_ids) == 0,
        'scoring_url': judge.get_scoring_url(),
        'qrcode_url': f'/api/judge/{judge.token}/qrcode/',
        'created_at': judge.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    }


def parse_category_id_list(value):
    if value in (None, ''):
        return []
    if not isinstance(value, list):
        raise ValueError('项目授权必须是数组')

    category_ids = []
    for item in value:
        try:
            category_id = int(item)
        except (TypeError, ValueError):
            raise ValueError('项目授权包含无效ID')
        if category_id <= 0:
            raise ValueError('项目授权包含无效ID')
        category_ids.append(category_id)
    return sorted(set(category_ids))


def apply_judge_category_filter(queryset, token):
    if not token:
        return queryset, None, None
    try:
        judge = Judge.objects.prefetch_related('allowed_categories').get(
            token=token,
            is_active=True,
        )
    except (Judge.DoesNotExist, ValidationError, ValueError):
        return queryset, None, error_response('无效的评委链接', 404)

    allowed_category_ids = get_judge_allowed_category_ids(judge)
    if allowed_category_ids:
        queryset = queryset.filter(id__in=allowed_category_ids)
    return queryset, judge, None


def format_extreme_rule_text(exclude_extreme_scores, exclude_lowest_count=1, exclude_highest_count=1):
    if not exclude_extreme_scores:
        return '统计时保留全部分数'

    lowest_count = max(0, to_int(exclude_lowest_count, 1))
    highest_count = max(0, to_int(exclude_highest_count, 1))
    total_drop_count = lowest_count + highest_count
    if total_drop_count <= 0:
        return '统计时保留全部分数'

    parts = []
    if lowest_count:
        parts.append(f'{lowest_count} 个最低分')
    if highest_count:
        parts.append(f'{highest_count} 个最高分')
    return f'统计时去掉{"、".join(parts)}（评分数需大于 {total_drop_count} 时生效）'


def format_score_rule_text(config):
    if config.score_value_type == 'integer':
        value_type_text = '整数'
    elif config.score_value_type == 'decimal':
        value_type_text = '小数（最多两位）'
    else:
        value_type_text = '整数或小数（最多两位）'
    duplicate_text = '允许重复打分' if config.allow_duplicate_scores else '不允许重复打分'
    extreme_text = format_extreme_rule_text(
        config.exclude_extreme_scores,
        config.exclude_lowest_count,
        config.exclude_highest_count,
    )
    return f'合法打分：{value_type_text}；{duplicate_text}；打分范围：{format_score_value(config.score_min)}-{format_score_value(config.score_max)}；{extreme_text}'


def format_category_rule_text(category, config):
    """格式化类别的统计规则文本"""
    mode = get_category_scoring_mode(category)
    if mode == 'vote':
        vote_total_count, vote_select_count = get_category_vote_params(category)
        return f'投票模式：从 {vote_total_count} 人中选择 {vote_select_count} 人'
    else:
        # 分数模式，获取类别的实际参数
        (
            score_min,
            score_max,
            score_value_type,
            allow_duplicate_scores,
            exclude_extreme_scores,
            exclude_lowest_count,
            exclude_highest_count,
        ) = get_category_score_params(category)
        if score_value_type == 'integer':
            value_type_text = '整数'
        elif score_value_type == 'decimal':
            value_type_text = '小数（最多两位）'
        else:
            value_type_text = '整数或小数（最多两位）'
        duplicate_text = '允许重复打分' if allow_duplicate_scores else '不允许重复打分'
        extreme_text = format_extreme_rule_text(
            exclude_extreme_scores,
            exclude_lowest_count,
            exclude_highest_count,
        )
        return f'合法打分：{value_type_text}；{duplicate_text}；打分范围：{format_score_value(score_min)}-{format_score_value(score_max)}；{extreme_text}'


def serialize_site_config(config, include_private=False):
    data = {
        'site_name': config.site_name,
        'primary_color': config.primary_color,
        'score_min': format_score_value(config.score_min),
        'score_max': format_score_value(config.score_max),
        'score_value_type': config.score_value_type,
        'allow_duplicate_scores': config.allow_duplicate_scores,
        'allow_scoring': config.allow_scoring,
        'exclude_extreme_scores': config.exclude_extreme_scores,
        'exclude_lowest_count': config.exclude_lowest_count,
        'exclude_highest_count': config.exclude_highest_count,
        'background_image': config.background_image.url if config.background_image else None,
        'logo_image': config.logo_image.url if config.logo_image else None,
        'scoring_mode': config.scoring_mode,
        'vote_total_count': config.vote_total_count,
        'vote_select_count': config.vote_select_count,
    }
    if include_private:
        data.update({
            'admin_password': config.admin_password,
            'base_url': config.base_url,
            'admin_url': config.get_admin_url(),
            'participant_sheet_name': config.participant_sheet_name,
            'category_field_name': config.category_field_name,
            'participant_field_name': config.participant_field_name,
            'order_field_name': config.order_field_name,
            'college_field_name': config.college_field_name,
        })
    return data


def render_qr_png_bytes(url):
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')

    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    return buffer.getvalue()


def build_qr_response(url, filename):
    response = HttpResponse(render_qr_png_bytes(url), content_type='image/png')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


def sanitize_filename_component(value):
    sanitized = re.sub(r'[\\/:*?"<>|]+', '_', normalize_text(value))
    sanitized = re.sub(r'\s+', '_', sanitized).strip('._')
    return sanitized or 'judge'


def build_judge_qrcode_filename(pattern, judge, index, site_name):
    template = normalize_text(pattern) or '{index}_{judge_name}'
    filename = template.format(
        index=index,
        judge_id=judge.id,
        judge_name=judge.name,
        judge_display_name=judge_display_name(judge),
        site_name=site_name,
        token=judge.token,
    )
    return sanitize_filename_component(filename)


def apply_score_rule(
    raw_scores,
    exclude_extreme_scores=False,
    exclude_lowest_count=1,
    exclude_highest_count=1,
):
    scores = list(raw_scores)
    lowest_count = max(0, to_int(exclude_lowest_count, 1))
    highest_count = max(0, to_int(exclude_highest_count, 1))
    total_drop_count = lowest_count + highest_count
    if exclude_extreme_scores and total_drop_count > 0 and len(scores) > total_drop_count:
        ordered = sorted(scores)
        high_start = len(ordered) - highest_count if highest_count else len(ordered)
        return ordered[lowest_count:high_start], ordered[:lowest_count], ordered[high_start:]
    return scores, [], []


def calculate_participant_statistics(
    participant,
    scores,
    exclude_extreme_scores=False,
    exclude_lowest_count=1,
    exclude_highest_count=1,
):
    if not scores:
        return None

    effective_scores, dropped_lows, dropped_highs = apply_score_rule(
        scores,
        exclude_extreme_scores,
        exclude_lowest_count,
        exclude_highest_count,
    )
    rule_applied = len(dropped_lows) > 0 or len(dropped_highs) > 0
    raw_total = quantize_score(sum(scores, Decimal('0')))
    raw_average = quantize_score(raw_total / Decimal(len(scores))) if scores else Decimal('0')
    total = quantize_score(sum(effective_scores, Decimal('0')))
    average = quantize_score(total / Decimal(len(effective_scores))) if effective_scores else Decimal('0')

    return {
        'participant_id': participant.id,
        'participant_name': participant.name,
        'college': participant.college,
        'raw_scores': [format_score_value(item) for item in scores],
        'total': format_score_value(total),
        'average': format_score_value(average),
        'count': len(scores),
        'effective_count': len(effective_scores),
        'raw_total': format_score_value(raw_total),
        'raw_average': format_score_value(raw_average),
        'dropped_low': format_score_value(dropped_lows[0]) if dropped_lows else None,
        'dropped_high': format_score_value(dropped_highs[-1]) if dropped_highs else None,
        'dropped_lows': [format_score_value(item) for item in dropped_lows],
        'dropped_highs': [format_score_value(item) for item in dropped_highs],
        'dropped_low_count': len(dropped_lows),
        'dropped_high_count': len(dropped_highs),
        'rule_applied': rule_applied,
    }


def _verify_admin(request):
    """验证管理员密码（Issue 8: 优先从 POST body 和 Header 读取，不再从 GET 参数读取）"""
    password = None
    # 优先从请求头读取
    password = request.META.get('HTTP_X_ADMIN_PASSWORD')
    # 其次从 POST 表单读取
    if not password:
        password = request.POST.get('password')
    # 最后从 JSON body 读取
    if not password:
        try:
            body = json.loads(request.body)
            password = body.get('password')
        except Exception:
            pass
    # 兼容 URL 参数读取（前端一些 POST 请求会将密码放在 URL 中 ?password=xxx）
    if not password:
        password = request.GET.get('password')
    config = SiteConfig.get_config()
    return password == config.admin_password


# ==================== 站点配置 ====================

@require_GET
def get_site_config(request):
    """获取站点公开配置"""
    config = SiteConfig.get_config()
    return json_response(serialize_site_config(config))


@require_GET
def get_admin_config(request):
    """获取管理员可编辑配置"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)
    config = SiteConfig.get_config()
    return json_response(serialize_site_config(config, include_private=True))


@csrf_exempt
@require_POST
def update_admin_config(request):
    """更新管理员配置，支持表单和JSON"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()
    payload = {}
    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return error_response('无效的JSON格式')
    else:
        payload = request.POST

    config.site_name = normalize_text(payload.get('site_name')) or config.site_name
    config.base_url = normalize_text(payload.get('base_url')) or config.base_url
    config.primary_color = normalize_text(payload.get('primary_color')) or config.primary_color
    config.score_min = normalize_decimal(payload.get('score_min'), config.score_min)
    config.score_max = normalize_decimal(payload.get('score_max'), config.score_max)
    score_value_type = normalize_text(payload.get('score_value_type')) or config.score_value_type
    if score_value_type in {'integer', 'decimal', 'integer_decimal'}:
        config.score_value_type = score_value_type
    config.allow_duplicate_scores = parse_bool(payload.get('allow_duplicate_scores'), config.allow_duplicate_scores)
    config.allow_scoring = parse_bool(payload.get('allow_scoring'), config.allow_scoring)
    config.exclude_extreme_scores = parse_bool(payload.get('exclude_extreme_scores'), config.exclude_extreme_scores)
    config.exclude_lowest_count = to_non_negative_int(
        payload.get('exclude_lowest_count'),
        config.exclude_lowest_count,
    )
    config.exclude_highest_count = to_non_negative_int(
        payload.get('exclude_highest_count'),
        config.exclude_highest_count,
    )
    
    scoring_mode = normalize_text(payload.get('scoring_mode'))
    if scoring_mode in ['score', 'vote']:
        config.scoring_mode = scoring_mode
    
    config.vote_total_count = to_int(payload.get('vote_total_count'), config.vote_total_count)
    config.vote_select_count = to_int(payload.get('vote_select_count'), config.vote_select_count)

    new_admin_password = normalize_text(payload.get('admin_password'))
    if new_admin_password:
        config.admin_password = new_admin_password

    config.participant_sheet_name = normalize_text(payload.get('participant_sheet_name')) or config.participant_sheet_name
    config.category_field_name = normalize_text(payload.get('category_field_name')) or config.category_field_name
    config.participant_field_name = normalize_text(payload.get('participant_field_name')) or config.participant_field_name
    config.order_field_name = normalize_text(payload.get('order_field_name')) or config.order_field_name

    if config.score_min >= config.score_max:
        return error_response('最高分必须大于最低分')

    if parse_bool(payload.get('clear_background'), False) and config.background_image:
        config.background_image.delete(save=False)
        config.background_image = None

    if request.FILES.get('background_image'):
        config.background_image = request.FILES['background_image']

    if request.FILES.get('logo_image'):
        config.logo_image = request.FILES['logo_image']

    config.save()
    return json_response({
        'success': True,
        'message': '配置已更新',
        'config': serialize_site_config(config, include_private=True),
    })


# ==================== 评委认证 ====================

def build_judge_submission_state(judge):
    """构建评委已提交记录的服务端快照。"""
    submitted_score_category_ids = set()
    submitted_scores = defaultdict(dict)
    judge_scores = (
        Score.objects.filter(judge=judge)
        .select_related('participant')
        .order_by(
            'participant__category__order',
            'participant__order',
            'participant_id',
        )
    )
    for score in judge_scores:
        category_id = score.participant.category_id
        submitted_score_category_ids.add(category_id)
        submitted_scores[category_id][score.participant_id] = format_score_value(score.score)

    submitted_vote_category_ids = set()
    submitted_votes = defaultdict(list)
    judge_votes = (
        Vote.objects.filter(judge=judge)
        .select_related('participant', 'category')
        .order_by(
            'category__order',
            'category_id',
            'vote_order',
            'participant__order',
            'participant_id',
        )
    )
    for vote in judge_votes:
        submitted_vote_category_ids.add(vote.category_id)
        submitted_votes[vote.category_id].append({
            'participant_id': vote.participant_id,
            'vote_order': vote.vote_order,
        })

    submitted_category_ids = sorted(
        submitted_score_category_ids | submitted_vote_category_ids
    )

    return {
        'submitted_categories': submitted_category_ids,
        'submitted_scores': {
            category_id: dict(category_scores)
            for category_id, category_scores in submitted_scores.items()
        },
        'submitted_votes': {
            category_id: list(category_votes)
            for category_id, category_votes in submitted_votes.items()
        },
    }


@require_GET
def judge_auth(request, token):
    """通过token验证评委身份"""
    try:
        judge = Judge.objects.prefetch_related('allowed_categories').get(
            token=token,
            is_active=True,
        )
    except Judge.DoesNotExist:
        return error_response('无效的评委链接', 404)

    submission_state = build_judge_submission_state(judge)
    allowed_category_ids = get_judge_allowed_category_ids(judge)
    
    # 获取该评委可参与类别的打分模式信息。未配置授权时默认可参与全部类别。
    categories = Category.objects.all()
    if allowed_category_ids:
        categories = categories.filter(id__in=allowed_category_ids)
    category_modes = {}
    for cat in categories:
        mode = get_category_scoring_mode(cat)
        total, select = get_category_vote_params(cat)
        (
            score_min,
            score_max,
            score_value_type,
            allow_duplicate,
            exclude_extreme,
            exclude_lowest_count,
            exclude_highest_count,
        ) = get_category_score_params(cat)
        category_modes[cat.id] = {
            'mode': mode,
            'vote_total_count': total,
            'vote_select_count': select,
            'score_min': format_score_value(score_min),
            'score_max': format_score_value(score_max),
            'score_value_type': score_value_type,
            'allow_duplicate_scores': allow_duplicate,
            'exclude_extreme_scores': exclude_extreme,
            'exclude_lowest_count': exclude_lowest_count,
            'exclude_highest_count': exclude_highest_count,
        }

    return json_response({
        'judge_id': judge.id,
        'judge_name': judge_display_name(judge),
        'allowed_category_ids': allowed_category_ids,
        'all_categories_allowed': len(allowed_category_ids) == 0,
        **submission_state,
        'category_modes': category_modes,
    })


# ==================== 获取数据 ====================

@require_GET
def get_categories(request):
    """获取所有比赛类别"""
    categories = Category.objects.prefetch_related('participants').all()
    categories, _, filter_error = apply_judge_category_filter(
        categories,
        normalize_text(request.GET.get('token')),
    )
    if filter_error:
        return filter_error
    result = []
    for cat in categories:
        result.append({
            'id': cat.id,
            'name': cat.name,
            'order': cat.order,
            'description': cat.description,
            'participant_count': cat.participants.count(),
            'scoring_mode': cat.scoring_mode,
            'vote_total_count': cat.vote_total_count,
            'vote_select_count': cat.vote_select_count,
            'score_min': format_score_value(cat.score_min),
            'score_max': format_score_value(cat.score_max),
            'score_value_type': cat.score_value_type,
            'allow_duplicate_scores': cat.allow_duplicate_scores,
            'exclude_extreme_scores': cat.exclude_extreme_scores,
            'exclude_lowest_count': cat.exclude_lowest_count,
            'exclude_highest_count': cat.exclude_highest_count,
        })
    return json_response({'categories': result})


@require_GET
def get_participants(request):
    """获取参赛选手列表，可按类别筛选"""
    category_id = request.GET.get('category_id')
    qs = Participant.objects.select_related('category').all()
    category_qs = Category.objects.all()
    category_qs, _, filter_error = apply_judge_category_filter(
        category_qs,
        normalize_text(request.GET.get('token')),
    )
    if filter_error:
        return filter_error
    qs = qs.filter(category_id__in=category_qs.values('id'))
    if category_id:
        qs = qs.filter(category_id=category_id)

    result = []
    for p in qs:
        result.append({
            'id': p.id,
            'name': p.name,
            'category_id': p.category_id,
            'category_name': p.category.name,
            'order': p.order,
            'description': p.description,
            'photo': p.photo.url if p.photo else None,
            'college': p.college,
        })
    return json_response({'participants': result})


# ==================== 评分操作 ====================

@csrf_exempt
@require_POST
@rate_limit(max_calls=10, period=60)
def submit_scores(request):
    """提交评分"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    token = data.get('token')
    category_id = data.get('category_id')
    scores = data.get('scores')

    if not token or not category_id or not scores:
        return error_response('缺少必要数据: token, category_id, scores')

    try:
        judge = Judge.objects.prefetch_related('allowed_categories').get(
            token=token,
            is_active=True,
        )
    except (Judge.DoesNotExist, ValidationError, ValueError):
        return error_response('无效的评委令牌', 403)

    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return error_response('类别不存在')

    if not judge_can_access_category(judge, category):
        return error_response(f'您没有权限提交【{category.name}】的评分', 403)

    has_submitted = Score.objects.filter(
        judge=judge,
        participant__category=category
    ).exists()
    if has_submitted:
        return error_response(f'您已经提交过【{category.name}】的评分，不可重复提交', 409)

    category_participants = set(
        Participant.objects.filter(category=category).values_list('id', flat=True)
    )

    # 获取类别级别的分数参数
    score_min, score_max, score_value_type, allow_duplicate_scores, _, _, _ = get_category_score_params(category)

    score_objects = []
    normalized_score_values = []
    for item in scores:
        pid = item.get('participant_id')
        score_val = item.get('score')

        if pid is None or score_val in (None, ''):
            return error_response('每条评分必须包含 participant_id 和 score')

        if pid not in category_participants:
            return error_response(f'参赛者ID {pid} 不属于类别【{category.name}】')

        try:
            decimal_score = quantize_score(Decimal(str(score_val).strip()))
        except (InvalidOperation, ValueError, TypeError):
            return error_response('分数格式无效')

        if score_value_type == 'integer' and decimal_score != decimal_score.to_integral_value():
            return error_response('当前只允许输入整数分')

        if score_value_type == 'decimal' and decimal_score == decimal_score.to_integral_value():
            return error_response('当前只允许输入小数分')

        if decimal_score < score_min or decimal_score > score_max:
            return error_response(f'分数必须在 {format_score_value(score_min)}-{format_score_value(score_max)} 之间')

        normalized_score_values.append(str(decimal_score))
        score_objects.append(Score(judge=judge, participant_id=pid, score=decimal_score))

    if len(score_objects) != len(category_participants):
        return error_response(f'需要为该类别的所有 {len(category_participants)} 名选手评分')

    if len({obj.participant_id for obj in score_objects}) != len(score_objects):
        return error_response('同一选手不能重复提交分数')

    if not allow_duplicate_scores:
        duplicates = [score for score, count in Counter(normalized_score_values).items() if count > 1]
        if duplicates:
            repeated = ', '.join(duplicates)
            return error_response(f'当前规则不允许重复打分，重复分数：{repeated}')

    Score.objects.bulk_create(score_objects)
    score_event_bus.notify()
    return json_response({
        'success': True,
        'message': f'成功提交【{category.name}】的评分',
        **build_judge_submission_state(judge),
    })


@csrf_exempt
@require_POST
@rate_limit(max_calls=10, period=60)
def submit_votes(request):
    """提交投票"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    token = data.get('token')
    category_id = data.get('category_id')
    votes = data.get('votes')

    if not token or not category_id or votes is None:
        return error_response('缺少必要数据: token, category_id, votes')

    try:
        judge = Judge.objects.prefetch_related('allowed_categories').get(
            token=token,
            is_active=True,
        )
    except (Judge.DoesNotExist, ValidationError, ValueError):
        return error_response('无效的评委令牌', 403)

    config = SiteConfig.get_config()
    if not config.allow_scoring:
        return error_response('评分已关闭，请联系管理员', 403)

    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return error_response('类别不存在')

    if not judge_can_access_category(judge, category):
        return error_response(f'您没有权限提交【{category.name}】的投票', 403)

    # 检查是否已经提交过
    has_submitted = Vote.objects.filter(
        judge=judge,
        category=category
    ).exists()
    if has_submitted:
        return error_response(f'您已经提交过【{category.name}】的投票，不可重复提交', 409)

    # 获取该类别的投票参数
    mode = get_category_scoring_mode(category)
    if mode != 'vote':
        return error_response(f'该类别当前不是投票模式', 400)
    
    vote_total_count, vote_select_count = get_category_vote_params(category)
    
    # 验证投票数量
    if len(votes) != vote_select_count:
        return error_response(f'需要选择 {vote_select_count} 名选手')

    # 获取该类别所有选手
    category_participants = set(
        Participant.objects.filter(category=category).values_list('id', flat=True)
    )

    # 验证选手
    vote_objects = []
    used_pids = set()
    for idx, item in enumerate(votes):
        pid = item.get('participant_id') if isinstance(item, dict) else item
        
        if pid is None:
            return error_response('每条投票必须包含 participant_id')

        if pid not in category_participants:
            return error_response(f'参赛者ID {pid} 不属于类别【{category.name}】')
        
        if pid in used_pids:
            return error_response(f'不能重复选择同一选手')
        
        used_pids.add(pid)
        vote_objects.append(Vote(
            judge=judge,
            category=category,
            participant_id=pid,
            vote_order=idx
        ))

    Vote.objects.bulk_create(vote_objects)
    score_event_bus.notify()
    return json_response({
        'success': True,
        'message': f'成功提交【{category.name}】的投票',
        **build_judge_submission_state(judge),
    })


# ==================== 管理员功能 ====================

@csrf_exempt
def verify_admin(request):
    if request.method == 'POST':
        if _verify_admin(request):
            return json_response({'success': True})
        return error_response('密码错误', 403)
    return error_response('方法不允许', 405)


@require_GET
def get_all_scores(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    result = build_scores_data()
    return json_response(result)


@require_GET
def export_excel(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    categories = Category.objects.prefetch_related('participants').all()
    judges = list(Judge.objects.filter(is_active=True).order_by('id'))
    all_scores = list(Score.objects.select_related('judge', 'participant', 'participant__category').all())
    all_votes = list(Vote.objects.select_related('judge', 'participant', 'category').all())

    scores_by_judge_participant = {(score.judge_id, score.participant_id): score.score for score in all_scores}
    scores_by_participant = defaultdict(list)
    for score in all_scores:
        scores_by_participant[score.participant_id].append(score.score)

    votes_by_participant = defaultdict(int)
    for vote in all_votes:
        votes_by_participant[vote.participant_id] += 1

    if not categories.exists():
        ws = wb.create_sheet(title='无数据')
        ws['A1'] = '暂无评分数据'
    else:
        for category in categories:
            participants = list(category.participants.all())
            mode = get_category_scoring_mode(category)
            ws = wb.create_sheet(title=category.name[:31])
            ws.append(['活动标题', config.site_name])

            if mode == 'vote':
                # 投票模式
                vote_total_count, vote_select_count = get_category_vote_params(category)
                total_votes = sum(votes_by_participant.get(p.id, 0) for p in participants)
                ws.append(['统计规则', f'投票模式：从 {vote_total_count} 人中选择 {vote_select_count} 人'])
                ws.append([])
                ws.append(['排名', '选手序号', '选手', '学院', '票数', '得票率'])

                # 计算投票统计
                vote_stats = []
                for participant in participants:
                    vote_count = votes_by_participant.get(participant.id, 0)
                    vote_rate = (vote_count / total_votes * 100) if total_votes > 0 else 0
                    vote_stats.append({
                        'participant_id': participant.id,
                        'participant': participant,
                        'vote_count': vote_count,
                        'vote_rate': vote_rate,
                    })
                # 按得票数降序排序
                vote_stats.sort(key=lambda item: (item['vote_count'], item['participant'].name), reverse=True)

                for index, stat in enumerate(vote_stats, start=1):
                    ws.append([
                        index,
                        stat['participant'].order,
                        stat['participant'].name,
                        stat['participant'].college or '',
                        stat['vote_count'],
                        f"{stat['vote_rate']:.2f}%",
                    ])
            else:
                # 打分模式
                ws.append(['统计规则', format_category_rule_text(category, config)])
                ws.append([])
                ws.append([
                    '排名', '选手序号', '选手', '学院', '统计总分', '统计平均分',
                    '原始总分', '原始平均分', '评委数', '统计计分数',
                    '去掉最低分', '去掉最高分', '是否去掉极值',
                ])

                ranking_stats = []
                # 获取类别的统计规则设置
                (
                    _,
                    _,
                    _,
                    _,
                    category_exclude_extreme,
                    exclude_lowest_count,
                    exclude_highest_count,
                ) = get_category_score_params(category)
                for participant in participants:
                    stat = calculate_participant_statistics(
                        participant,
                        scores_by_participant.get(participant.id, []),
                        category_exclude_extreme,
                        exclude_lowest_count,
                        exclude_highest_count,
                    )
                    if stat:
                        ranking_stats.append(stat)
                ranking_stats.sort(key=lambda item: (item['total'], item['average'], item['participant_name']), reverse=True)

                for index, stat in enumerate(ranking_stats, start=1):
                    participant = next((p for p in participants if p.id == stat['participant_id']), None)
                    ws.append([
                        index,
                        participant.order if participant else '',
                        stat['participant_name'],
                        stat['college'],
                        stat['total'],
                        stat['average'],
                        stat['raw_total'],
                        stat['raw_average'],
                        stat['count'],
                        stat['effective_count'],
                        format_score_list(stat['dropped_lows']),
                        format_score_list(stat['dropped_highs']),
                        '是' if stat['rule_applied'] else '否',
                    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="scores_export.xlsx"'
    return response


@require_GET
def export_score_details(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()
    wb = Workbook()

    categories = Category.objects.prefetch_related('participants').all()
    judges = list(Judge.objects.filter(is_active=True).order_by('order', 'id'))
    all_scores = list(Score.objects.select_related('judge', 'participant', 'participant__category').all())
    all_votes = list(Vote.objects.select_related('judge', 'participant', 'category').all())

    scores_by_judge_participant = {(score.judge_id, score.participant_id): score.score for score in all_scores}
    votes_by_judge_category = defaultdict(list)
    for vote in all_votes:
        votes_by_judge_category[(vote.judge_id, vote.category_id)].append({
            'participant_id': vote.participant_id,
            'vote_order': vote.vote_order,
        })

    # 为每个类别创建单独的sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for category in categories:
        mode = get_category_scoring_mode(category)
        ws = wb.create_sheet(title=category.name[:31])
        participants = list(category.participants.all().order_by('order', 'id'))

        if mode == 'vote':
            # 投票模式明细
            vote_total_count, vote_select_count = get_category_vote_params(category)
            ws.append(['投票模式：从 {} 人中选择 {} 人'.format(vote_total_count, vote_select_count)])
            ws.append([])

            headers = ['评委', '第1选择', '第2选择', '第3选择']
            if vote_select_count > 3:
                for i in range(4, vote_select_count + 1):
                    headers.append(f'第{i}选择')
            ws.append(headers)

            for judge in judges:
                judge_votes = votes_by_judge_category.get((judge.id, category.id), [])
                # 按vote_order排序
                judge_votes.sort(key=lambda x: x['vote_order'])
                # 使用与打分模式相同的评委标签生成逻辑
                scoring_url = judge.get_scoring_url()
                parts = scoring_url.split('/')
                label = parts[-1]
                label_part = str(label)[:len(label)//2]
                row = [label_part]
                vote_names = []
                for vote in judge_votes:
                    participant = next((p for p in participants if p.id == vote['participant_id']), None)
                    if participant:
                        vote_names.append(participant.name)
                # 填充选择
                for i in range(vote_select_count):
                    row.append(vote_names[i] if i < len(vote_names) else '')
                ws.append(row)
        else:
            # 打分模式明细
            headers = ['选手序号', '选手', '学院']
            judge_column_labels = []
            for judge in judges:
                scoring_url = judge.get_scoring_url()
                parts = scoring_url.split('/')
                label = parts[-1]
                label_part = str(label)[:len(label)//2]
                judge_column_labels.append(label_part)
            headers.extend(judge_column_labels)
            ws.append(headers)

            for participant in participants:
                row = [
                    participant.order,
                    participant.name,
                    participant.college or ''
                ]
                for judge in judges:
                    score = scores_by_judge_participant.get((judge.id, participant.id))
                    row.append(format_score_value(score) if score is not None else '')
                ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="score_details.xlsx"'
    return response


@require_GET
def export_participants(request):
    """导出选手信息表格（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()
    wb = Workbook()
    ws = wb.active
    ws.title = config.participant_sheet_name[:31]
    
    ws.append([
        config.category_field_name,
        config.participant_field_name,
        config.order_field_name,
        config.college_field_name,
    ])
    
    participants = Participant.objects.select_related('category').order_by('category__order', 'order', 'id')
    for participant in participants:
        ws.append([
            participant.category.name if participant.category else '',
            participant.name,
            participant.order,
            participant.college or '',
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="participants_export.xlsx"'
    return response


@require_GET
def download_import_template(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()
    wb = Workbook()
    participant_ws = wb.active
    participant_ws.title = config.participant_sheet_name[:31]
    participant_ws.append([
        config.category_field_name,
        config.participant_field_name,
        config.order_field_name,
        config.college_field_name,
    ])
    participant_ws.append(['才艺组', '选手A', 1, '计算机学院'])
    participant_ws.append(['才艺组', '选手B', 2, '外国语学院'])
    participant_ws.append(['演讲组', '选手C', 1, '经济学院'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="import_template.xlsx"'
    return response


@csrf_exempt
@require_POST
def clear_scores(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    clear_password = request.POST.get('clear_password') or request.GET.get('clear_password')
    if not clear_password:
        try:
            body = json.loads(request.body)
            clear_password = body.get('clear_password')
        except Exception:
            clear_password = None

    clear_password = normalize_text(clear_password)
    if not clear_password:
        return error_response('请输入清空评分密码', 400)

    if clear_password != ADMIN_CLEAR_PASSWORD:
        return error_response('清空评分密码错误', 403)

    score_count = Score.objects.count()
    vote_count = Vote.objects.count()
    Score.objects.all().delete()
    Vote.objects.all().delete()
    total_count = score_count + vote_count
    score_event_bus.notify()
    return json_response({'success': True, 'message': f'已清空 {total_count} 条记录（{score_count} 条评分，{vote_count} 条投票）'})


# ==================== 评委管理 ====================

@require_GET
def get_judges(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    judges = Judge.objects.prefetch_related('allowed_categories').all().order_by('order', 'id')
    result = [serialize_judge(judge) for judge in judges]
    return json_response({'judges': result})


@csrf_exempt
@require_POST
def batch_create_judges(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    names = data.get('names', [])
    count = data.get('count', 0)

    created = []
    next_order = get_next_judge_order()
    if names:
        for name in names:
            normalized_name = normalize_text(name)
            if not normalized_name:
                continue
            judge = Judge.objects.create(order=next_order, name=normalized_name)
            created.append({'id': judge.id, 'order': judge.order, 'name': judge.name, 'token': str(judge.token)})
            next_order += 1
    elif count > 0:
        for _ in range(count):
            judge = Judge.objects.create(order=next_order, name=f'评委{next_order}')
            created.append({'id': judge.id, 'order': judge.order, 'name': judge.name, 'token': str(judge.token)})
            next_order += 1

    return json_response({'success': True, 'created': created, 'count': len(created)})


@csrf_exempt
@require_POST
def update_judge(request, judge_id):
    """更新评委信息（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    try:
        judge = Judge.objects.get(id=judge_id)
    except Judge.DoesNotExist:
        return error_response('评委不存在', 404)

    name = normalize_text(data.get('name', ''))
    if not name:
        return error_response('评委名称不能为空')

    order = to_int(data.get('order'), judge.order)
    if order <= 0:
        return error_response('评委序号必须大于 0')

    allowed_categories = None
    if 'allowed_category_ids' in data:
        try:
            allowed_category_ids = parse_category_id_list(data.get('allowed_category_ids'))
        except ValueError as exc:
            return error_response(str(exc))
        allowed_categories = list(Category.objects.filter(id__in=allowed_category_ids))
        if len(allowed_categories) != len(allowed_category_ids):
            return error_response('项目授权包含不存在的类别')

    duplicate = Judge.objects.exclude(id=judge.id).filter(order=order).exists()
    if duplicate:
        return error_response('该评委序号已被使用')

    judge.name = name
    judge.order = order
    judge.save()
    if allowed_categories is not None:
        judge.allowed_categories.set(allowed_categories)

    return json_response({
        'success': True,
        'message': '评委信息已更新',
        'judge': serialize_judge(judge),
    })


@csrf_exempt
@require_POST
def delete_judge(request, judge_id):
    """删除评委（管理员）- Issue 15: 删除前检查关联评分数据"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        judge = Judge.objects.get(id=judge_id)
    except Judge.DoesNotExist:
        return error_response('评委不存在', 404)

    # Issue 15: 检查关联数据
    score_count = Score.objects.filter(judge=judge).count()
    vote_count = Vote.objects.filter(judge=judge).count()

    # 检查是否强制删除
    force = False
    try:
        body = json.loads(request.body)
        force = body.get('force', False)
    except Exception:
        pass

    if (score_count > 0 or vote_count > 0) and not force:
        return error_response(
            f'评委 {judge_display_name(judge)} 有 {score_count} 条评分和 {vote_count} 条投票记录，'
            f'删除后这些数据将丢失。如确认删除，请传入 force=true',
            409
        )

    try:
        judge_name = judge_display_name(judge)
        judge.delete()
        return json_response({
            'success': True,
            'message': f'评委 {judge_name} 已删除（同时删除了 {score_count} 条评分和 {vote_count} 条投票记录）'
        })
    except Exception as e:
        logger.exception('删除评委失败')
        return error_response(f'删除失败: {str(e)}', 500)


# ==================== 数据导入 ====================

def get_header_index_map(sheet):
    headers = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    normalized_map = {normalize_header(header): index for index, header in enumerate(headers) if header}
    return headers, normalized_map


@csrf_exempt
def import_data(request):
    """导入数据（管理员）- 仅导入选手数据，支持 Excel 和 JSON"""
    if request.method != 'POST':
        return error_response('方法不允许', 405)
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    config = SiteConfig.get_config()

    uploaded_file = request.FILES.get('file')
    if uploaded_file:
        try:
            workbook = load_workbook(uploaded_file, data_only=True)
        except Exception as exc:
            logger.exception('导入Excel失败')
            return error_response(f'无法读取Excel文件: {exc}')

        participant_sheet_name = config.participant_sheet_name
        participant_sheet = workbook[participant_sheet_name] if participant_sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]

        _, participant_header_map = get_header_index_map(participant_sheet)
        category_idx = participant_header_map.get(normalize_header(config.category_field_name))
        participant_idx = participant_header_map.get(normalize_header(config.participant_field_name))
        order_idx = participant_header_map.get(normalize_header(config.order_field_name))
        college_idx = participant_header_map.get(normalize_header(config.college_field_name))

        if category_idx is None or participant_idx is None:
            return error_response(
                f'Excel缺少必要字段：{config.category_field_name}、{config.participant_field_name}'
            )

        created_categories = 0
        created_participants = 0
        updated_participants = 0
        category_order_map = {}
        participant_sequence_map = defaultdict(int)

        for row in participant_sheet.iter_rows(min_row=2, values_only=True):
            values = list(row)
            category_name = normalize_text(values[category_idx] if len(values) > category_idx else '')
            participant_name = normalize_text(values[participant_idx] if len(values) > participant_idx else '')
            if not category_name and not participant_name:
                continue
            if not category_name or not participant_name:
                continue

            if category_name not in category_order_map:
                category_order_map[category_name] = len(category_order_map) + 1

            category, category_created = Category.objects.get_or_create(
                name=category_name,
                defaults={'order': category_order_map[category_name]}
            )
            if category_created:
                created_categories += 1

            participant_sequence_map[category_name] += 1
            participant_order = participant_sequence_map[category_name]
            if order_idx is not None and len(values) > order_idx:
                participant_order = to_int(values[order_idx], participant_order)
            
            college = ''
            if college_idx is not None and len(values) > college_idx:
                college = normalize_text(values[college_idx])

            # 检查是否已存在相同姓名、类别和序号的选手
            try:
                participant = Participant.objects.get(
                    name=participant_name,
                    category=category,
                    order=participant_order
                )
                if college:
                    participant.college = college
                    participant.save()
                participant_created = False
            except Participant.DoesNotExist:
                participant = Participant.objects.create(
                    name=participant_name,
                    category=category,
                    order=participant_order,
                    college=college
                )
                participant_created = True
                created_participants += 1

        return json_response({
            'success': True,
            'message': (
                f'Excel导入完成：新增{created_categories}个类别，新增{created_participants}名选手'
            ),
        })

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('请上传Excel文件，或提交有效的JSON格式数据')

    # 处理手动导入的选手列表
    participants_data = data.get('participants', [])
    if participants_data:
        created_categories = 0
        created_participants = 0
        updated_participants = 0
        category_order_map = {}
        participant_sequence_map = defaultdict(int)

        for p_data in participants_data:
            category_name = normalize_text(p_data.get('category', ''))
            participant_name = normalize_text(p_data.get('name', ''))
            participant_order = p_data.get('order')
            college = normalize_text(p_data.get('college', ''))

            if not category_name or not participant_name:
                continue

            if category_name not in category_order_map:
                category_order_map[category_name] = len(category_order_map) + 1

            category, category_created = Category.objects.get_or_create(
                name=category_name,
                defaults={'order': category_order_map[category_name]}
            )
            if category_created:
                created_categories += 1

            participant_sequence_map[category_name] += 1
            if participant_order is None:
                participant_order = participant_sequence_map[category_name]

            # 检查是否已存在相同姓名、类别和序号的选手
            try:
                participant = Participant.objects.get(
                    name=participant_name,
                    category=category,
                    order=participant_order
                )
                if college:
                    participant.college = college
                    participant.save()
                participant_created = False
            except Participant.DoesNotExist:
                participant = Participant.objects.create(
                    name=participant_name,
                    category=category,
                    order=participant_order,
                    college=college
                )
                participant_created = True
                created_participants += 1

        return json_response({
            'success': True,
            'message': (
                f'手动导入完成：新增{created_categories}个类别，新增{created_participants}名选手'
            ),
        })

    # 处理原有的 categories 格式
    categories_data = data.get('categories', [])
    created_categories = 0
    created_participants = 0

    for cat_data in categories_data:
        cat_name = normalize_text(cat_data.get('name', ''))
        if not cat_name:
            continue
        category, category_created = Category.objects.get_or_create(
            name=cat_name,
            defaults={
                'order': cat_data.get('order', 0),
                'description': cat_data.get('description', ''),
            }
        )
        if category_created:
            created_categories += 1

        for idx, p_data in enumerate(cat_data.get('participants', []), start=1):
            participant_name = p_data if isinstance(p_data, str) else p_data.get('name', '')
            participant_name = normalize_text(participant_name)
            participant_order = p_data.get('order', idx) if isinstance(p_data, dict) else idx
            college = ''
            if isinstance(p_data, dict):
                college = normalize_text(p_data.get('college', ''))
            if not participant_name:
                continue
            # 检查是否已存在相同姓名、类别和序号的选手
            try:
                participant = Participant.objects.get(
                    name=participant_name,
                    category=category,
                    order=participant_order
                )
                if college:
                    participant.college = college
                    participant.save()
                participant_created = False
            except Participant.DoesNotExist:
                participant = Participant.objects.create(
                    name=participant_name,
                    category=category,
                    order=participant_order,
                    description=p_data.get('description', '') if isinstance(p_data, dict) else '',
                    college=college
                )
                participant_created = True
                created_participants += 1

    return json_response({
        'success': True,
        'message': f'导入完成：新增{created_categories}个类别，新增{created_participants}名选手',
    })


# ==================== 二维码 ====================

@require_GET
def generate_qrcode(request, token):
    try:
        judge = Judge.objects.get(token=token)
    except Judge.DoesNotExist:
        return error_response('评委不存在', 404)
    safe_name = judge_display_name(judge).replace(' ', '_')
    return build_qr_response(judge.get_scoring_url(), f'judge_{safe_name}_qr.png')


@require_GET
def export_all_judge_qrcodes(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    judges = list(Judge.objects.all().order_by('id'))
    if not judges:
        return error_response('暂无评委可导出', 404)

    config = SiteConfig.get_config()
    pattern = normalize_text(request.GET.get('pattern')) or '{index}_{judge_name}'
    used_names = set()
    try:
        build_judge_qrcode_filename(pattern, judges[0], 1, config.site_name)
    except KeyError as exc:
        return error_response(f'命名规则包含不支持的占位符：{exc.args[0]}')
    except Exception:
        return error_response('二维码命名规则格式无效')
    output = io.BytesIO()

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as archive:
        for index, judge in enumerate(judges, start=1):
            base_name = build_judge_qrcode_filename(pattern, judge, index, config.site_name)
            filename = f'{base_name}.png'
            duplicate_index = 2
            while filename in used_names:
                filename = f'{base_name}_{duplicate_index}.png'
                duplicate_index += 1
            used_names.add(filename)
            archive.writestr(filename, render_qr_png_bytes(judge.get_scoring_url()))

    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="judge_qrcodes.zip"'
    return response


@require_GET
def generate_admin_qrcode(request):
    if not _verify_admin(request):
        return error_response('权限不足', 403)
    config = SiteConfig.get_config()
    return build_qr_response(config.get_admin_url(), 'admin_qr.png')


# ==================== 选手管理 ====================

@require_GET
def get_admin_participants(request):
    """获取所有选手列表（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    participants = Participant.objects.select_related('category').all()
    result = []
    for p in participants:
        result.append({
            'id': p.id,
            'name': p.name,
            'category_id': p.category_id,
            'category_name': p.category.name,
            'order': p.order,
            'college': p.college,
        })
    return json_response({'participants': result})


@csrf_exempt
@require_POST
def delete_participant(request, participant_id):
    """删除选手（管理员）- Issue 15: 删除前检查关联评分数据"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        participant = Participant.objects.get(id=participant_id)
    except Participant.DoesNotExist:
        return error_response('选手不存在', 404)

    # Issue 15: 检查关联数据
    score_count = Score.objects.filter(participant=participant).count()
    vote_count = Vote.objects.filter(participant=participant).count()

    # 检查是否强制删除
    force = False
    try:
        body = json.loads(request.body)
        force = body.get('force', False)
    except Exception:
        pass

    if (score_count > 0 or vote_count > 0) and not force:
        return error_response(
            f'选手 {participant.name} 有 {score_count} 条评分和 {vote_count} 条投票记录，'
            f'删除后这些数据将丢失。如确认删除，请传入 force=true',
            409
        )

    try:
        participant_name = participant.name
        participant.delete()
        return json_response({
            'success': True,
            'message': f'选手 {participant_name} 已删除（同时删除了 {score_count} 条评分和 {vote_count} 条投票记录）'
        })
    except Exception as e:
        logger.exception('删除选手失败')
        return error_response(f'删除失败: {str(e)}', 500)


@csrf_exempt
@require_POST
def clear_participants(request):
    """清空所有选手（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    clear_password = request.POST.get('clear_password') or request.GET.get('clear_password')
    if not clear_password:
        try:
            body = json.loads(request.body)
            clear_password = body.get('clear_password')
        except Exception:
            clear_password = None

    clear_password = normalize_text(clear_password)
    if not clear_password:
        return error_response('请输入清空密码', 400)

    if clear_password != ADMIN_CLEAR_PASSWORD:
        return error_response('清空密码错误', 403)

    count = Participant.objects.count()
    Participant.objects.all().delete()
    return json_response({'success': True, 'message': f'已清空 {count} 名选手'})


@csrf_exempt
@require_POST
def clear_judges(request):
    """清空所有评委（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    clear_password = request.POST.get('clear_password') or request.GET.get('clear_password')
    if not clear_password:
        try:
            body = json.loads(request.body)
            clear_password = body.get('clear_password')
        except Exception:
            clear_password = None

    clear_password = normalize_text(clear_password)
    if not clear_password:
        return error_response('请输入清空密码', 400)

    if clear_password != ADMIN_CLEAR_PASSWORD:
        return error_response('清空密码错误', 403)

    count = Judge.objects.count()
    Judge.objects.all().delete()
    return json_response({'success': True, 'message': f'已清空 {count} 名评委'})


@csrf_exempt
@require_POST
def clear_categories(request):
    """清空所有类别（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    clear_password = request.POST.get('clear_password') or request.GET.get('clear_password')
    if not clear_password:
        try:
            body = json.loads(request.body)
            clear_password = body.get('clear_password')
        except Exception:
            clear_password = None

    clear_password = normalize_text(clear_password)
    if not clear_password:
        return error_response('请输入清空密码', 400)

    if clear_password != ADMIN_CLEAR_PASSWORD:
        return error_response('清空密码错误', 403)

    count = Category.objects.count()
    Category.objects.all().delete()
    return json_response({'success': True, 'message': f'已清空 {count} 个类别'})


@csrf_exempt
@require_POST
def update_participant(request, participant_id):
    """更新选手信息（管理员），支持自动创建新类别"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    try:
        participant = Participant.objects.get(id=participant_id)
    except Participant.DoesNotExist:
        return error_response('选手不存在', 404)

    name = normalize_text(data.get('name', ''))
    if not name:
        return error_response('选手姓名不能为空')

    category_id = data.get('category_id')
    category_name = normalize_text(data.get('category_name', ''))

    category = None
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            pass

    if not category and category_name:
        category, _ = Category.objects.get_or_create(name=category_name, defaults={'order': get_next_category_order()})

    if not category:
        return error_response('请指定有效类别')

    participant.name = name
    participant.category = category
    participant.order = to_int(data.get('order'), participant.order)
    participant.college = normalize_text(data.get('college', ''))
    participant.save()

    return json_response({
        'success': True,
        'message': '选手信息已更新',
        'participant': {
            'id': participant.id,
            'name': participant.name,
            'category_id': participant.category_id,
            'category_name': participant.category.name,
            'order': participant.order,
            'college': participant.college,
        }
    })


@csrf_exempt
@require_POST
def create_participant(request):
    """创建选手（管理员），支持自动创建新类别"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    name = normalize_text(data.get('name', ''))
    if not name:
        return error_response('选手姓名不能为空')

    category_id = data.get('category_id')
    category_name = normalize_text(data.get('category_name', ''))

    category = None
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            pass

    if not category and category_name:
        category, _ = Category.objects.get_or_create(name=category_name, defaults={'order': get_next_category_order()})

    if not category:
        return error_response('请指定有效类别')

    order = to_int(data.get('order'))
    if order <= 0:
        max_order = Participant.objects.filter(category=category).aggregate(models.Max('order'))['order__max']
        order = (max_order or 0) + 1

    participant = Participant.objects.create(
        name=name,
        category=category,
        order=order,
        college=normalize_text(data.get('college', ''))
    )

    return json_response({
        'success': True,
        'message': '选手已创建',
        'participant': {
            'id': participant.id,
            'name': participant.name,
            'category_id': participant.category_id,
            'category_name': participant.category.name,
            'order': participant.order,
            'college': participant.college,
        }
    })


@csrf_exempt
@require_POST
def create_category(request):
    """创建类别（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    name = normalize_text(data.get('name', ''))
    if not name:
        return error_response('类别名称不能为空')

    if Category.objects.filter(name=name).exists():
        return error_response('该类别已存在')

    order = to_int(data.get('order'), 0)
    if order <= 0:
        order = get_next_category_order()
    
    scoring_mode = normalize_text(data.get('scoring_mode', 'default'))
    if scoring_mode not in ['default', 'score', 'vote']:
        scoring_mode = 'default'
    
    vote_total_count = to_int(data.get('vote_total_count')) if 'vote_total_count' in data else None
    vote_select_count = to_int(data.get('vote_select_count')) if 'vote_select_count' in data else None
    
    # 分数模式相关字段
    score_min = normalize_decimal(data.get('score_min')) if 'score_min' in data else None
    score_max = normalize_decimal(data.get('score_max')) if 'score_max' in data else None
    score_value_type = normalize_text(data.get('score_value_type')) if 'score_value_type' in data else None
    allow_duplicate_scores = parse_bool(data.get('allow_duplicate_scores')) if 'allow_duplicate_scores' in data else None
    exclude_extreme_scores = parse_bool(data.get('exclude_extreme_scores')) if 'exclude_extreme_scores' in data else None
    exclude_lowest_count = to_non_negative_int(data.get('exclude_lowest_count'), None) if 'exclude_lowest_count' in data else None
    exclude_highest_count = to_non_negative_int(data.get('exclude_highest_count'), None) if 'exclude_highest_count' in data else None

    category = Category.objects.create(
        name=name,
        order=order,
        description=normalize_text(data.get('description', '')),
        scoring_mode=scoring_mode,
        vote_total_count=vote_total_count,
        vote_select_count=vote_select_count,
        score_min=score_min,
        score_max=score_max,
        score_value_type=score_value_type,
        allow_duplicate_scores=allow_duplicate_scores,
        exclude_extreme_scores=exclude_extreme_scores,
        exclude_lowest_count=exclude_lowest_count,
        exclude_highest_count=exclude_highest_count,
    )

    return json_response({
        'success': True,
        'message': '类别已创建',
        'category': {
            'id': category.id,
            'name': category.name,
            'order': category.order,
            'description': category.description,
            'participant_count': 0,
            'scoring_mode': category.scoring_mode,
            'vote_total_count': category.vote_total_count,
            'vote_select_count': category.vote_select_count,
            'score_min': format_score_value(category.score_min),
            'score_max': format_score_value(category.score_max),
            'score_value_type': category.score_value_type,
            'allow_duplicate_scores': category.allow_duplicate_scores,
            'exclude_extreme_scores': category.exclude_extreme_scores,
            'exclude_lowest_count': category.exclude_lowest_count,
            'exclude_highest_count': category.exclude_highest_count,
        }
    })


@csrf_exempt
@require_POST
def update_category(request, category_id):
    """更新类别（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return error_response('无效的JSON格式')

    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return error_response('类别不存在', 404)

    name = normalize_text(data.get('name', ''))
    if name and name != category.name:
        if Category.objects.filter(name=name).exists():
            return error_response('该类别名称已被使用')
        category.name = name

    if 'order' in data:
        category.order = to_int(data.get('order'), category.order)

    if 'description' in data:
        category.description = normalize_text(data.get('description', ''))
    
    if 'scoring_mode' in data:
        scoring_mode = normalize_text(data.get('scoring_mode'))
        if scoring_mode in ['default', 'score', 'vote']:
            category.scoring_mode = scoring_mode
    
    if 'vote_total_count' in data:
        val = data.get('vote_total_count')
        category.vote_total_count = to_int(val) if val is not None else None
    
    if 'vote_select_count' in data:
        val = data.get('vote_select_count')
        category.vote_select_count = to_int(val) if val is not None else None
    
    # 分数模式相关字段
    if 'score_min' in data:
        val = data.get('score_min')
        category.score_min = normalize_decimal(val) if val is not None else None
    
    if 'score_max' in data:
        val = data.get('score_max')
        category.score_max = normalize_decimal(val) if val is not None else None
    
    if 'score_value_type' in data:
        val = normalize_text(data.get('score_value_type'))
        category.score_value_type = val if val and val in ['integer', 'decimal', 'integer_decimal'] else None
    
    if 'allow_duplicate_scores' in data:
        category.allow_duplicate_scores = parse_bool(data.get('allow_duplicate_scores')) if 'allow_duplicate_scores' in data else None
    
    if 'exclude_extreme_scores' in data:
        category.exclude_extreme_scores = parse_bool(data.get('exclude_extreme_scores')) if 'exclude_extreme_scores' in data else None

    if 'exclude_lowest_count' in data:
        val = data.get('exclude_lowest_count')
        category.exclude_lowest_count = to_non_negative_int(val, None) if val is not None else None

    if 'exclude_highest_count' in data:
        val = data.get('exclude_highest_count')
        category.exclude_highest_count = to_non_negative_int(val, None) if val is not None else None

    category.save()

    return json_response({
        'success': True,
        'message': '类别已更新',
        'category': {
            'id': category.id,
            'name': category.name,
            'order': category.order,
            'description': category.description,
            'participant_count': category.participants.count(),
            'scoring_mode': category.scoring_mode,
            'vote_total_count': category.vote_total_count,
            'vote_select_count': category.vote_select_count,
            'score_min': format_score_value(category.score_min),
            'score_max': format_score_value(category.score_max),
            'score_value_type': category.score_value_type,
            'allow_duplicate_scores': category.allow_duplicate_scores,
            'exclude_extreme_scores': category.exclude_extreme_scores,
            'exclude_lowest_count': category.exclude_lowest_count,
            'exclude_highest_count': category.exclude_highest_count,
        }
    })


@csrf_exempt
@require_POST
def delete_category(request, category_id):
    """删除类别（管理员）"""
    if not _verify_admin(request):
        return error_response('权限不足', 403)

    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return error_response('类别不存在', 404)

    category_name = category.name
    category.delete()

    return json_response({'success': True, 'message': f'类别 {category_name} 已删除'})
